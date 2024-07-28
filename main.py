import time
from pdfrw import PdfReader, PdfWriter
import pygame
import openpyxl
import fitz
import tkinter as tk
from tkinter import filedialog
from PyPDF2 import PdfWriter, PdfReader
import io
import sys
import re
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter
from tkinter import messagebox
import hashlib
import os

OPTIONS = [
    "OZON",
    "WILDBERRIES",
    "CDEK",
    "YANDEX",
    "BOXBERRY"
]

X_WB_POSITION = 6
Y_WB_POSITION = 4

X_OZON_POSITION = 15
Y_OZON_POSITION = 85

SPLIT_SIZE = 150

root = tk.Tk()
root.title("Добавление значений на PDF")
root.geometry("640x480")  # Установка размера окна

result = tk.StringVar()
pdf_file_path = tk.StringVar()
excel_file_path = tk.StringVar()
output_pdf_file_path = tk.StringVar()
output_excel_file_path = tk.StringVar()
output_folder_path = tk.StringVar()  # Добавляем переменную для хранения пути к папке
variable = tk.StringVar(root)
variable.set(OPTIONS[0])
entry_size = tk.Entry(root)
entry_y = tk.Entry(root)
entry_x = tk.Entry(root)

def resource_path(relative):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(relative)


def optimizePDF(path: str, new_path: str):
    document = fitz.open(path)

    new_document = fitz.open()

    for page_num in range(document.page_count):
        page = document.load_page(page_num)
        new_page = new_document.new_page(width=page.rect.width, height=page.rect.height)

        new_page.show_pdf_page(new_page.rect, document, page_num)

    new_document.save(new_path, deflate=True)
    document.close()
    new_document.close()


def browse_pdf_file():
    pdf_file_path.set(filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")]))
    output_pdf_file_path.set(pdf_file_path.get().split('/')[-1])


def browse_excel_file():
    excel_file_path.set(filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")]))
    output_excel_file_path.set(excel_file_path.get().split('/')[-1])


def browse_output_folder():
    output_folder = filedialog.askdirectory()
    output_folder_path.set(output_folder)


def split_pdf(input_path):
    document = fitz.open(input_path)
    total_pages = document.page_count
    parts = []

    for i in range(0, total_pages, SPLIT_SIZE):
        part = fitz.open()
        part.insert_pdf(document, from_page=i, to_page=min(i + SPLIT_SIZE - 1, total_pages - 1))
        part_path = f"part_{i // SPLIT_SIZE}.pdf"
        part.save(part_path)
        part.close()
        parts.append(part_path)

    document.close()
    return parts


def merge_pdfs(output_path, pdf_paths):
    output = fitz.open()

    for pdf_path in pdf_paths:
        pdf_document = fitz.open(pdf_path)
        output.insert_pdf(pdf_document)
        pdf_document.close()

    output.save(output_path)
    output.close()


def delete_pdf_files(file_list):
    for file_path in file_list:
        if file_path.endswith('.pdf'):
            delete_file(file_path, 5)


def delete_file(file_path, retries):
    flag = False
    for _ in range(retries):
        try:
            if not flag:
                time.sleep(3)
                doc = fitz.open(file_path)
                doc.close()
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
                flag = True
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")
            result.set("Something went wrong while deleting file, check directory and delete temp files.")
            time.sleep(2)


def sort_pdf(pdf_path, output_pdf_path, elems: set):
    key = str()
    doc = fitz.open(pdf_path)
    pages = []
    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)
        text = page.get_text()
        for elem in elems:
            match = re.search(elem, text)
            if match:
                key = match.group(0)
        pages.append((key, page_num, page))

    pages.sort(key=lambda x: x[0])

    new_doc = fitz.open()
    for _, _, page in pages:
        new_doc.insert_pdf(doc, from_page=page.number, to_page=page.number)

    new_doc.save(output_pdf_path)
    new_doc.close()
    doc.close()


def start_processing():
    parts = split_pdf(pdf_file_path.get())
    new_parts = []
    output_path = f'{output_folder_path.get()}/edited_{variable.get()}.pdf'
    output_opt_path = f'{output_folder_path.get()}/opt_edited_{variable.get()}.pdf'
    output_sort_opt_path = f'{output_folder_path.get()}/sort_opt_edited_{variable.get()}.pdf'

    elems = set()

    if variable.get() == OPTIONS[0] or variable.get() == OPTIONS[2] or variable.get() == OPTIONS[4]:
        for part in parts:
            start_processing_cdek_ozon_boxberry(part, elems)
            new_parts.append('opt_new_' + part)
            os.remove('new_' + part)

        merge_pdfs(output_path, new_parts)
    elif variable.get() == OPTIONS[1]:
        start_processing_wb()
    else:
        start_processing_yandex()

    optimizePDF(output_path, output_opt_path)
    sort_pdf(output_opt_path, output_sort_opt_path, elems)
    result.set("Calculation complete. Check the file!!!")
    delete_pdf_files(parts)
    delete_pdf_files(new_parts)


def start_processing_cdek_ozon_boxberry(pdf_path, elems: set):
    path = "new_" + pdf_path

    with open(path, "wb") as output_stream:
        excel_path = excel_file_path.get()

        with open(pdf_path, "rb") as pdf_file:
            pdf_document = PdfReader(pdf_file)
            excel_workbook = openpyxl.load_workbook(excel_path)
            excel_sheet = excel_workbook.active
            output = PdfWriter()

            for pdf_page_number, pdf_page in enumerate(pdf_document.pages):
                pdf_all_text = pdf_page.extract_text().replace(' ', '')
                pdf_text = re.search(r'\d+-\d+-\d+', pdf_all_text).group(0)
                matching_values = []

                for row in excel_sheet.iter_rows(min_row=2, min_col=2, max_col=12, values_only=True):
                    if any(cell is not None for cell in row):
                        excel_value_to_match = row[0] if row[0] is not None else ""
                        excel_value_from_col10 = row[9] if row[9] is not None else ""
                        excel_value_from_col12 = row[10] if row[10] is not None else ""
                        excel_value_from_col12_str = str(excel_value_from_col12)
                        excel_value_from_col10_str = str(excel_value_from_col10)
                        if excel_value_to_match == pdf_text:
                            matching_values.append((excel_value_from_col10_str, excel_value_from_col12_str))

                if matching_values:
                    packet = io.BytesIO()
                    can = canvas.Canvas(packet, pagesize=letter)
                    can.setFont('FreeSans', int(entry_size.get()))
                    text_object = can.beginText()
                    text_object.setTextOrigin(int(entry_x.get()), int(entry_y.get()))
                    text_object.setFont('FreeSans', int(entry_size.get()))

                    combined_values = []
                    for value_10, value_12 in matching_values:
                        combined_values.append(f"{value_10}, {value_12}")
                    combined_values_text = " ; ".join(combined_values)

                    elems.add(combined_values_text)

                    max_line_length = 30
                    lines = [combined_values_text[i:i + max_line_length] for i in
                             range(0, len(combined_values_text), max_line_length)]

                    for line in lines:
                        text_object.textLine(line)

                    can.drawText(text_object)
                    can.save()
                    new_pdf = PdfReader(packet)
                    pdf_page.merge_page(new_pdf.pages[0])
                    output.add_page(pdf_page)

            output.write(output_stream)
            output.close()

    optimizePDF(path, 'opt_' + path)
    return elems


def start_processing_yandex():
    pdf_path = pdf_file_path.get()
    path = f'{output_folder_path.get()}/non_opt_edited_yandex.pdf'
    output_stream = open(path, "wb")
    excel_path = excel_file_path.get()
    pdf_document = PdfReader(open(pdf_path, "rb"))
    excel_workbook = openpyxl.load_workbook(excel_path)
    excel_sheet = excel_workbook.active
    output = PdfWriter()
    for pdf_page_number, pdf_page in enumerate(pdf_document.pages):
        pdf_text = pdf_page.extract_text().replace('\n', '').split()[0]
        matching_values = []

        for row in excel_sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            if any(cell is not None for cell in row):
                excel_value_to_match = str(row[0]) if row[0] is not None else ""
                name = str(row[1]) if row[1] is not None else ""
                quantity = str(row[2]) if row[2] is not None else ""
                if excel_value_to_match == pdf_text:
                    matching_values.append((name, quantity))
        if matching_values:
            packet = io.BytesIO()
            can = canvas.Canvas(packet, pagesize=letter)
            can.setFont('FreeSans', int(entry_size.get()))

            text_object = can.beginText()
            text_object.setTextOrigin(int(entry_x.get()), int(entry_y.get()))
            text_object.setFont('FreeSans', int(entry_size.get()))

            combined_values = []

            for value_10, value_12 in matching_values:
                combined_values.append(f"{value_10}, {value_12}")
            combined_values_text = " ; ".join(combined_values)

            max_line_length = 30
            lines = [combined_values_text[i:i + max_line_length] for i in
                     range(0, len(combined_values_text), max_line_length)]

            for line in lines:
                text_object.textLine(line)

            can.rotate(90)
            can.drawText(text_object)
            can.save()
            new_pdf = PdfReader(packet)
            pdf_page.merge_page(new_pdf.pages[0])
            output.add_page(pdf_page)

    output.write(output_stream)
    output_stream.close()
    optimizePDF(path, f'{output_folder_path.get()}/edited_{variable.get()}.pdf')
    os.remove(path)


def start_processing_wb():
    pdf_path = pdf_file_path.get()
    path = f'{output_folder_path.get()}/non_opt_edited_wb.pdf'
    output_stream = open(path, "wb")
    excel_path = excel_file_path.get()
    pdf_document = PdfReader(open(pdf_path, "rb"))
    excel_workbook = openpyxl.load_workbook(excel_path)
    excel_sheet = excel_workbook.active
    output = PdfWriter()
    for pdf_page_number, pdf_page in enumerate(pdf_document.pages):
        pdf_text = pdf_page.extract_text()
        pdf_text = pdf_text.replace("\n", "")
        pdf_text = re.sub(r'[a-zA-Zа-яА-Я]', '', pdf_text)
        matching_value = ""
        for row in excel_sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            if any(cell is not None for cell in row):
                excel_value_to_match = row[1].replace(" ", '') if row[1] is not None else ""
                excel_value_from_col0 = row[0] if row[0] is not None else ""
                excel_value_from_col0_str = str(excel_value_from_col0)

                if excel_value_to_match == pdf_text:
                    matching_value = excel_value_from_col0_str
                    break
        if len(matching_value) != 0:
            packet = io.BytesIO()
            can = canvas.Canvas(packet, pagesize=letter)
            can.setFont('FreeSans', int(entry_size.get()))
            can.rotate(90)
            can.drawString(int(entry_x.get()), -int(entry_y.get()), matching_value)
            can.save()
            new_pdf = PdfReader(packet)
            pdf_page.merge_page(new_pdf.pages[0])
            output.add_page(pdf_page)
    output.write(output_stream)
    output_stream.close()
    optimizePDF(path, f'{output_folder_path.get()}/edited_{variable.get()}.pdf')
    os.remove(path)


# Генерация хеша из лицензионного ключа
def generate_hash(key):
    return hashlib.sha256(key.encode()).hexdigest()


# Проверка лицензионного ключа
def check_key(key):
    correct_hash = "f6f5301aa729ae10124cf5dd6ff5431b5183e2d617e076c49d5940c46c490254"
    return generate_hash(key) == correct_hash


# Функция для открытия главного окна приложения
def open_main_window():
    option_label = tk.Label(root, text="Выберите сервис")
    option_label.pack()

    w = tk.OptionMenu(root, variable, *OPTIONS)
    w.pack()

    pdf_label = tk.Label(root, text="Выберите PDF файл:")
    pdf_label.pack()

    pdf_button = tk.Button(root, text="Обзор", command=browse_pdf_file)
    pdf_button.pack()

    pdf_folder_label = tk.Label(root, textvariable=output_pdf_file_path)
    pdf_folder_label.pack()

    excel_label = tk.Label(root, text="Выберите Excel файл:")
    excel_label.pack()

    excel_button = tk.Button(root, text="Обзор", command=browse_excel_file)
    excel_button.pack()

    excel_folder_label = tk.Label(root, textvariable=output_excel_file_path)
    excel_folder_label.pack()

    x_label = tk.Label(root, text="X координата:")
    x_label.pack()

    entry_x.pack()

    y_label = tk.Label(root, text="Y координата:")
    y_label.pack()

    entry_y.pack()

    size_label = tk.Label(root, text="Размер шрифта: ")
    size_label.pack()

    entry_size.pack()

    pygame.init()
    font = resource_path(os.path.join('Folder', 'FreeSans.ttf'))

    output_folder_button = tk.Button(root, text="Выбрать папку для сохранения", command=browse_output_folder)
    output_folder_button.pack()

    output_folder_label = tk.Label(root, textvariable=output_folder_path)
    output_folder_label.pack()

    process_button = tk.Button(root, text="Начать", command=start_processing)
    process_button.pack()

    result_label = tk.Label(root, textvariable=result)
    result_label.pack()

    root.mainloop()


# Функция для проверки лицензионного ключа при запуске приложения
def verify_license():
    license_window = tk.Tk()
    license_window.title("Ввод лицензионного ключа")

    tk.Label(license_window, text="Введите лицензионный ключ:").pack()
    license_entry = tk.Entry(license_window)
    license_entry.pack()

    def on_submit():
        key = license_entry.get()
        if check_key(key):
            # Сохранение лицензионного ключа в файл
            with open("license.key", "w") as f:
                f.write(generate_hash(key))
            license_window.destroy()
            open_main_window()
        else:
            messagebox.showerror("Ошибка", "Неверный лицензионный ключ!")

    submit_button = tk.Button(license_window, text="Подтвердить", command=on_submit)
    submit_button.pack()

    license_window.mainloop()


# Проверка наличия и правильности лицензионного ключа в файле
def check_saved_key():
    if os.path.exists("license.key"):
        with open("license.key", "r") as f:
            saved_key = f.read().strip()
        correct_hash = "f6f5301aa729ae10124cf5dd6ff5431b5183e2d617e076c49d5940c46c490254"
        if saved_key == correct_hash:
            return True
    return False


# Основная логика
if check_saved_key():
    open_main_window()
else:
    verify_license()

pdfmetrics.registerFont(TTFont('FreeSans', 'FreeSans.ttf', 'CP1251'))



